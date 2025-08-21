from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path


def leer_valores_horizontales(ws, fila, col_inicio=4):
    valores = []
    col = col_inicio
    while True:
        val = ws.cell(row=fila, column=col).value
        if val is None:
            break
        try:
            valores.append(int(val))
        except (ValueError, TypeError):
            pass
        col += 1
    return valores


def leer_titulos(ws, fila, cantidad):
    return [
        str(ws.cell(row=fila, column=col).value).strip()
        for col in range(4, 4 + cantidad)
    ]


def escribir_encabezado(ws, fila, texto):
    negrita = Font(bold=True)
    fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = negrita
    c.fill = fill
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)


def escribir_tabla(ws, fila_inicio, encabezados, datos):
    negrita = Font(bold=True)
    for i, encabezado in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_inicio, column=i, value=encabezado)
        c.font = negrita

    for j, fila_dato in enumerate(datos, start=fila_inicio + 1):
        for i, dato in enumerate(fila_dato, start=1):
            ws.cell(row=j, column=i, value=dato)


def generar_metricas(path_excel: Path):
    wb = load_workbook(path_excel)
    ws_resumen = wb["Resumen"]

    # === Obtener datos ===
    cites = leer_valores_horizontales(ws_resumen, fila=10)
    views = leer_valores_horizontales(ws_resumen, fila=14)
    titulos = leer_titulos(ws_resumen, fila=3, cantidad=len(cites))

    total_articulos = len(cites)
    total_cites = sum(cites)
    total_views = sum(views)

    promedio_cites = total_cites / total_articulos if total_articulos else 0
    promedio_views = total_views / total_articulos if total_articulos else 0

    # === Crear hoja nueva ===
    if "Métricas" in wb.sheetnames:
        wb.remove(wb["Métricas"])
    ws_metrica = wb.create_sheet("Métricas")

    fila = 1
    escribir_encabezado(ws_metrica, fila, "Resumen de Métricas Generales")
    fila += 1

    resumen = [
        ["Total de artículos", total_articulos],
        ["Total de citas", total_cites],
        ["Total de views", total_views],
        ["Promedio de citas por artículo", round(promedio_cites, 2)],
        ["Promedio de views por artículo", round(promedio_views, 2)],
    ]
    for item in resumen:
        ws_metrica.cell(row=fila, column=1, value=item[0])
        ws_metrica.cell(row=fila, column=2, value=item[1])
        fila += 1

    # === Top 5 citados ===
    fila += 2
    escribir_encabezado(ws_metrica, fila, "Top 5 artículos más citados")
    fila += 1
    top_citados = sorted(zip(titulos, cites), key=lambda x: x[1], reverse=True)[:5]
    escribir_tabla(ws_metrica, fila, ["Título", "Citas"], top_citados)
    fila += len(top_citados) + 2

    # === Top 5 vistos ===
    escribir_encabezado(ws_metrica, fila, "Top 5 artículos más vistos")
    fila += 1
    top_vistos = sorted(zip(titulos, views), key=lambda x: x[1], reverse=True)[:5]
    escribir_tabla(ws_metrica, fila, ["Título", "Views"], top_vistos)
    fila += len(top_vistos) + 2

    # === Artículos sin citas ===
    sin_citas = [1 for c in cites if c == 0]
    cant_sin_citas = len(sin_citas)
    porcentaje_sin_citas = cant_sin_citas / total_articulos if total_articulos else 0

    escribir_encabezado(ws_metrica, fila, "Artículos sin citas")
    fila += 1
    ws_metrica.cell(row=fila, column=1, value="Cantidad").font = Font(bold=True)
    ws_metrica.cell(row=fila, column=2, value=cant_sin_citas)
    fila += 1
    ws_metrica.cell(row=fila, column=1, value="Porcentaje").font = Font(bold=True)
    ws_metrica.cell(row=fila, column=2, value=porcentaje_sin_citas)
    ws_metrica.cell(row=fila, column=2).number_format = '0.00%'

    wb.save(path_excel)
    print("--> Métricas generadas correctamente.")


if __name__ == "__main__":
    generar_metricas(Path("Resumen.xlsx"))
