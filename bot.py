from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import urlparse
import shutil
from FN_IEEE import obtener_titulo_ieee, obtener_location_ieee, obtener_metricas_ieee, obtener_cita_ieee
from FN_SPRINGER import obtener_titulo_springer, obtener_location_springer, obtener_metricas_springer, obtener_cita_springer
from FN_ACM import obtener_titulo_acm, obtener_location_acm, obtener_metricas_acm, obtener_cita_acm

#
ORIGINAL_XLSX = Path("ExcelModelo/ExcelModelo.xlsx")
COPIA_XLSX = Path("Resumen.xlsx")
EXCEL_URLS = Path("ListadoArticulos.xlsx")

# Mapeo las funciones según el origen de la URL
EXTRACTORES = {
    "ieeexplore.ieee.org": {
        "titulo": obtener_titulo_ieee,
        "ubicacion": obtener_location_ieee,
        "metricas": obtener_metricas_ieee,
        "cita": obtener_cita_ieee
    },
    "link.springer.com": {
        "titulo": obtener_titulo_springer,
        "ubicacion": obtener_location_springer,
        "metricas": obtener_metricas_springer,
        "cita": obtener_cita_springer
    },
    "dl.acm.org": {
        "titulo": obtener_titulo_acm,
        "ubicacion": obtener_location_acm,
        "metricas": obtener_metricas_acm,
        "cita": obtener_cita_acm
    }
}

#Leo desde un archivo
def leer_urls_desde_excel(ruta_excel: Path) -> list[str]:
    wb = load_workbook(ruta_excel)
    ws = wb.active
    urls = []

    # Suponemos que el título está en A1 y los datos arrancan desde A2
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        link = row[0]
        if link and isinstance(link, str):
            urls.append(link.strip())

    return urls

#Funciones Utiles
def copiar_excel(origen: Path, destino: Path):
    shutil.copy(origen, destino)
    print(f"Copiando Excel...")

def inicializar_driver():
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    # 🔇 Desactiva los logs molestos
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def obtener_urls_existentes(ws, fila=2, desde_columna=4):
    urls = set()
    col = desde_columna
    while ws.cell(row=fila, column=col).value:
        urls.add(ws.cell(row=fila, column=col).value.strip())
        col += 1
    return urls

def encontrar_columna_libre(ws, fila=1, desde_columna=4):
    col = desde_columna
    while ws.cell(row=fila, column=col).value:
        col += 1
    return col

def escribir_articulo_en_excel(ws, col, articulo):
    ws.cell(row=1, column=col, value=articulo["id"])
    ws.cell(row=2, column=col, value=articulo["link"])
    ws.cell(row=3, column=col, value=articulo["title"])
    ws.cell(row=4, column=col, value=articulo["author"])
    ws.cell(row=5, column=col, value=articulo["year"])
    ws.cell(row=7, column=col, value=articulo["fuente"])
    ws.cell(row=8, column=col, value=articulo["booktitle"])
    ws.cell(row=9, column=col, value=articulo["ubicacion"])
    ws.cell(row=10, column=col, value=articulo["cites_in"])
    ws.cell(row=11, column=col, value=articulo["keywords"])
    ws.cell(row=12, column=col, value=articulo["cita"])
    ws.cell(row=13, column=col, value=articulo["doi"])
    ws.cell(row=14, column=col, value=articulo["text_views"])

# Main:
if __name__ == "__main__":
    #Tomo links del excel:
    URLS = leer_urls_desde_excel(EXCEL_URLS)
    # abrir/crear Excel
    if COPIA_XLSX.exists():
        wb = load_workbook(COPIA_XLSX)
        print("Excel existente.")
    else:
        copiar_excel(ORIGINAL_XLSX, COPIA_XLSX)
        wb = load_workbook(COPIA_XLSX)
        print("Creando Excel")
    ws = wb["Resumen"]

    urls_existentes = obtener_urls_existentes(ws)

    for url in URLS:
        if url in urls_existentes:
            print(f"-URL Ya Incorporada: {url}")
            continue

        print(f"Procesando: {url}")

        dominio = urlparse(url).netloc
        extractor = EXTRACTORES.get(dominio)
        if "ieee" in dominio:
            fuente = "IEEE"
        elif "springer" in dominio:
            fuente = "Springer"
        elif "acm" in dominio:
            fuente = "ACM"
        else:
            fuente = "Desconocida"
        if extractor is None:
            print(f"!!-> No hay extractor definido para {dominio}")
            continue
        driver = inicializar_driver()
        titulo = extractor["titulo"](driver, url)
        datos_cita = extractor["cita"](driver)
        ubicacion = extractor["ubicacion"](driver)
        cites_in, text_views = extractor["metricas"](driver)

        col_libre = encontrar_columna_libre(ws)
        articulo = {
            "id": col_libre - 3,
            "link": url,
            "title": titulo,
            "cita": datos_cita["cita"],
            "author": datos_cita["author"],
            "booktitle": datos_cita["booktitle"],
            "year": int(datos_cita["year"]) if str(datos_cita["year"]).isdigit() else 0,
            "keywords": datos_cita["keywords"],
            "doi": datos_cita["doi"],
            "fuente": fuente,
            "ubicacion": ubicacion,
            "cites_in": cites_in,
            "text_views": text_views
        }

        escribir_articulo_en_excel(ws, col_libre, articulo)
        wb.save(COPIA_XLSX)
        driver.delete_all_cookies()
        driver.execute_script("window.localStorage.clear();")
        driver.execute_script("window.sessionStorage.clear();")
        driver.quit()        
    #driver.quit()
    print("Proceso Finalizado.")
