from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from collections import Counter
from pathlib import Path


def crear_hoja_graficos(wb, nombre="Graficos"):
    if nombre in wb.sheetnames:
        wb.remove(wb[nombre])
    return wb.create_sheet(nombre)


def obtener_valores(ws_origen, fila, desde_columna=4):
    valores = []
    col = desde_columna
    while ws_origen.cell(row=fila, column=col).value:
        valor = ws_origen.cell(row=fila, column=col).value
        if valor:
            valores.append(str(valor).strip())
        col += 1
    return valores


def escribir_titulo(ws, fila, texto):
    verde_titulo = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    negrita = Font(bold=True)

    c = ws.cell(row=fila, column=1, value=texto)
    c.font = negrita
    c.fill = verde_titulo
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)


def escribir_tabla(ws_destino, conteo, total, fila_inicio, titulo_col1):
    verde_oliva = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    negrita = Font(bold=True)

    # encabezados
    ws_destino.cell(row=fila_inicio, column=1, value=titulo_col1)
    ws_destino.cell(row=fila_inicio, column=2, value="Cantidad")
    ws_destino.cell(row=fila_inicio, column=3, value="Porcentaje")
    for col in range(1, 4):
        c = ws_destino.cell(row=fila_inicio, column=col)
        c.font = negrita
        c.fill = verde_oliva

    # datos
    for idx, (valor, cantidad) in enumerate(conteo.items(), start=fila_inicio+1):
        ws_destino.cell(row=idx, column=1, value=valor).font = negrita
        ws_destino.cell(row=idx, column=2, value=cantidad)

        porcentaje_cell = ws_destino.cell(row=idx, column=3,
                                          value=f"=B{idx}/B{len(conteo)+fila_inicio+1}")
        porcentaje_cell.number_format = '0%'

    # fila total
    fila_total = idx+1
    ws_destino.cell(row=fila_total, column=1, value="Total").font = negrita
    ws_destino.cell(row=fila_total, column=2, value=total)
    for col in range(1, 4):
        c = ws_destino.cell(row=fila_total, column=col)
        c.font = negrita
        c.fill = verde_oliva

    return fila_total


def agregar_grafico(ws_destino, fila_inicio, filas_datos, titulo, celda, mostrar_leyenda):
    chart = BarChart()
    chart.title = titulo
    print(titulo)
    print(mostrar_leyenda)
    chart.y_axis.delete = False
    if not mostrar_leyenda:
        print("x_axis")
        chart.x_axis.delete = False

    data = Reference(ws_destino, min_col=2, min_row=fila_inicio,
                     max_row=fila_inicio + filas_datos)
    cats = Reference(ws_destino, min_col=1, min_row=fila_inicio + 1,
                     max_row=fila_inicio + filas_datos)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    if not mostrar_leyenda:
        print("leyenda")
        chart.legend = None  
    

    ws_destino.add_chart(chart, celda)



def generar_tabla_y_grafico(path_excel: Path):
    wb = load_workbook(path_excel)
    we_resumen = wb["Resumen"]

    ws_graficos = crear_hoja_graficos(wb)

    # === AÑOS ===
    anios = obtener_valores(we_resumen, fila=5)
    total_anios = len(anios)
    conteo_anios = Counter(map(int, anios))
    conteo_anios = dict(sorted(conteo_anios.items(), reverse=True))

    fila_actual = 1
    escribir_titulo(ws_graficos, fila_actual, "ARTÍCULOS POR AÑO")
    fila_actual += 1

    fila_fin_anios = escribir_tabla(ws_graficos, conteo_anios, total_anios,
                                    fila_inicio=fila_actual, titulo_col1="Año")
    agregar_grafico(ws_graficos, fila_actual, len(conteo_anios),
                    "Cantidad de artículos por Año", "E2",False)

    # === PAISES ===
    fila_actual = fila_fin_anios + 20  # más espacio entre tablas
    escribir_titulo(ws_graficos, fila_actual, "ARTÍCULOS POR PAÍS")
    fila_actual += 1

    paises = obtener_valores(we_resumen, fila=9)
    total_paises = len(paises)
    conteo_paises = Counter(paises)
    conteo_paises = dict(sorted(conteo_paises.items(), key=lambda x: x[1], reverse=True))

    fila_fin_paises = escribir_tabla(ws_graficos, conteo_paises, total_paises,fila_inicio=fila_actual, titulo_col1="País")
    agregar_grafico(ws_graficos, fila_actual, len(conteo_paises),
                    "Cantidad de artículos por País", f"E{fila_actual}",False)

    # === ARTÍCULOS POR FUENTES ===
    fila_actual = fila_fin_paises + 20  # más espacio entre tablas
    escribir_titulo(ws_graficos, fila_actual, "ARTÍCULOS POR FUENTES")
    fila_actual += 1

    fuentes = obtener_valores(we_resumen, fila=7)
    total_fuentes = len(fuentes)
    conteo_fuentes = Counter(fuentes)
    conteo_fuentes = dict(sorted(conteo_fuentes.items(), key=lambda x: x[1], reverse=True))

    fila_fin_fuentes = escribir_tabla(ws_graficos, conteo_fuentes, total_fuentes,fila_inicio=fila_actual, titulo_col1="Fuente")
    agregar_grafico(ws_graficos, fila_actual, len(conteo_fuentes),
                    "Cantidad de artículos por Fuentes", f"E{fila_actual}",False)    
    
    # === KEYWORDS MÁS USADAS (TOP 5) ===
    fila_actual = fila_fin_fuentes + 20  # espacio desde el gráfico anterior
    escribir_titulo(ws_graficos, fila_actual, "TOP 5 KEYWORDS MÁS USADAS")
    fila_actual += 1

    # Obtener valores de fila 11 (keywords), separados por coma
    keywords_raw = obtener_valores(we_resumen, fila=11)

    # Dividir y limpiar cada keyword
    lista_keywords = []
    for celda in keywords_raw:
        if celda:  # evitar celdas vacías
            celda = celda.replace(";", ",")
            palabras = [k.strip().lower() for k in celda.split(",") if k.strip()]
            lista_keywords.extend(palabras)

    # Contar y ordenar
    conteo_keywords = Counter(lista_keywords)
    top5_keywords = dict(conteo_keywords.most_common(5))

    # Escribir tabla y gráfico
    fila_fin_keywords = escribir_tabla(ws_graficos, top5_keywords, total=len(top5_keywords), fila_inicio=fila_actual, titulo_col1="Keyword")
    agregar_grafico(ws_graficos, fila_actual, len(top5_keywords), "Top 5 Keywords más usadas", f"E{fila_actual}",True)


    wb.save(path_excel)
    print(f"->Graficos generados")


if __name__ == "__main__":
    generar_tabla_y_grafico(Path("Resumen.xlsx"))
